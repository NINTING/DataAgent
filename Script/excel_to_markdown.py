import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

import pandas as pd
import openpyxl
import tempfile
import os
from pathlib import Path
from tabulate import tabulate


def unmerge_cells(wb, sheet_name):
    """
    处理工作表中的合并单元格，将合并值填充到所有单元格
    
    Args:
        wb: openpyxl 工作簿对象
        sheet_name: 工作表名称
        
    Returns:
        wb: 处理后的工作簿对象
    """
    ws = wb[sheet_name]
    
    # 获取所有合并单元格范围
    merged_ranges = list(ws.merged_cells.ranges)
    
    # 遍历每个合并单元格范围
    for merged in merged_ranges:
        # 获取合并区域的左上角单元格的值
        value = ws.cell(merged.min_row, merged.min_col).value
        
        # 取消合并
        ws.unmerge_cells(str(merged))
        
        # 将值填充到合并区域的所有单元格
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                ws.cell(row, col).value = value
    
    return wb


def convert_sheet_to_markdown(excel_file, sheet_name, output_dir, add_row_col_numbers=True):
    """
    转换单个工作表为 Markdown 格式并保存
    
    Args:
        excel_file: Excel 文件路径
        sheet_name: 工作表名称
        output_dir: 输出目录路径
        add_row_col_numbers: 是否添加行列序号（类似 Excel 的 A,B,C 和 1,2,3）
        
    Returns:
        str: 生成的 Markdown 文件路径
    """
    # 创建输出目录（如果不存在）
    os.makedirs(output_dir, exist_ok=True)
    
    # 处理合并单元格
    wb = openpyxl.load_workbook(excel_file)
    wb = unmerge_cells(wb, sheet_name)
    
    # 使用 openpyxl 直接提取数据（避免 pandas 读取时丢失公式）
    ws = wb[sheet_name]
    
    # 获取工作表的所有数据
    data = []
    for row in ws.iter_rows():
        row_values = []
        for cell in row:
            # 处理单元格值：None 转为空字符串，其他值转为字符串
            value = cell.value if cell.value is not None else ''
            row_values.append(str(value))
        data.append(row_values)
    
    # 分离表头和数据行
    if len(data) > 0:
        headers = data[0]
        rows = data[1:] if len(data) > 1 else []
    else:
        headers = []
        rows = []
    
    # 添加行列序号
    if add_row_col_numbers and len(data) > 0:
        num_cols = len(data[0])
        
        # 生成列序号 A, B, C... Z, AA, AB...
        col_headers = []
        for i in range(num_cols):
            col_letter = ''
            num = i
            while num >= 0:
                col_letter = chr(num % 26 + ord('A')) + col_letter
                num = num // 26 - 1
                if num < 0:
                    break
            col_headers.append(col_letter)
        
        # 构建表头：第一列留空，其余是列序号
        new_headers = [''] + col_headers
        
        # 构建数据行：每行前面添加行号（1, 2, 3...）
        new_rows = []
        for idx, row in enumerate(rows, start=1):
            new_rows.append([str(idx)] + row)
        
        # 更新 headers 和 rows
        headers = new_headers
        rows = new_rows
    
    # 使用 tabulate 生成 Markdown 表格
    markdown_table = tabulate(rows, headers=headers, tablefmt='pipe')
    
    # 生成安全的文件名（替换空格和斜杠）
    safe_name = sheet_name.replace(' ', '_').replace('/', '_')
    output_file = os.path.join(output_dir, f'{safe_name}.md')
    
    # 写入 Markdown 文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(markdown_table)
    
    return output_file


def convert_excel_to_markdown(excel_file, output_dir, add_row_col_numbers=True):
    """
    转换 Excel 文件的所有工作表为 Markdown
    
    Args:
        excel_file: Excel 文件路径
        output_dir: 输出目录路径
        add_row_col_numbers: 是否添加行列序号（类似 Excel 的 A,B,C 和 1,2,3）
        
    Returns:
        list: 所有生成的 Markdown 文件路径列表
    """
    # 读取 Excel 文件，获取所有工作表名称
    xl = pd.ExcelFile(excel_file)
    sheet_names = xl.sheet_names
    
    output_files = []
    
    # 遍历每个工作表并转换
    for sheet_name in sheet_names:
        output_file = convert_sheet_to_markdown(excel_file, sheet_name, output_dir, add_row_col_numbers)
        output_files.append(output_file)
        print(f'已保存: {output_file}')
    
    return output_files


def main():
    """
    程序入口点
    """
    excel_file = 'ExcelData/Covid Dashboard.xlsx'
    output_dir = 'ExcelData/Structured_MD'
    add_row_col_numbers = True  # 是否添加行列序号
    
    output_files = convert_excel_to_markdown(excel_file, output_dir, add_row_col_numbers)
    
    print(f'\n所有工作表转换完成！共生成 {len(output_files)} 个 Markdown 文件。')


if __name__ == '__main__':
    main()
