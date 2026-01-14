import openpyxl
import pandas as pd
import tempfile
import os
import warnings
warnings.filterwarnings('ignore')

excel_file = 'ExcelData/Covid Dashboard.xlsx'
sheet_name = 'Sheet1'

# 查看原始文件中公式单元格的值
print("=== 原始文件中 B18 单元格 ===")
wb = openpyxl.load_workbook(excel_file)
ws = wb[sheet_name]
cell_b18 = ws['B18']
print(f"  value: {cell_b18.value}")
print(f"  type: {type(cell_b18.value)}")
print(f"  is formula: {isinstance(cell_b18.value, str) and cell_b18.value.startswith('=')}")

# 查看合并单元格处理后
print("\n=== 合并单元格处理 ===")
merged_ranges = list(ws.merged_cells.ranges)
print(f"  合并单元格数量: {len(merged_ranges)}")

# 模拟完整流程 - 创建新工作簿避免修改问题
print("\n=== 完整流程测试 ===")
wb2 = openpyxl.load_workbook(excel_file)
ws2 = wb2[sheet_name]

# 合并单元格处理
merged_ranges = list(ws2.merged_cells.ranges)
for merged in merged_ranges:
    value = ws2.cell(merged.min_row, merged.min_col).value
    ws2.unmerge_cells(str(merged))
    for row in range(merged.min_row, merged.max_row + 1):
        for col in range(merged.min_col, merged.max_col + 1):
            ws2.cell(row, col).value = value

print(f"  处理后 B18 值: {ws2['B18'].value}")
print(f"  B18 值类型: {type(ws2['B18'].value)}")

# 保存后 pandas 读取
with tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx') as tmp:
    wb2.save(tmp.name)
    tmp_path = tmp.name

df = pd.read_excel(tmp_path, sheet_name=sheet_name, header=None)
os.unlink(tmp_path)

print(f"\n=== pandas 读取结果 ===")
print(f"  B18 (行17, 列1) 值: {df.iloc[17, 1]}")
print(f"  B18 值类型: {type(df.iloc[17, 1])}")
